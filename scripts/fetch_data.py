#!/usr/bin/env python3
# scripts/fetch_data.py
# ────────────────────────────────────────────────────────────────────────
# SSOT 데이터(구글 시트 6탭 + Notion 아이템 DB) → data/*.json 추출 파이프라인.
#
# 원본:
#   - 시트: https://docs.google.com/spreadsheets/d/1iS4Lmjx32w0Mu527foFtc8pQn3pw6APcQYnvcKdVM9o/
#           탭: 베이스탐험카드 / 1차_확정카드 / 전투카드 / 몬스터 / 사냥감 / 빌딩 / 아이템마스터
#   - Notion 아이템 DB: 레거시(참조용 서사 보관). D-30부터 런타임 SSOT는 시트 '아이템마스터'.
#
# 시트 액세스 (D-25 전환):
#   primary — Google Sheets API (서비스 계정). 읽기 + 쓰기 둘 다 가능.
#     의존성: gspread, google-auth
#     키 경로 기본값: .secrets/sheets-sa.json
#     환경변수: GOOGLE_SHEETS_SA_KEY=/절대/경로.json  (있으면 덮어씀)
#   fallback — 공개 xlsx export URL (읽기 전용).
#     API 초기화가 실패하거나 --no-api 플래그 지정 시 사용.
#
# 산출물(덮어쓰기):
#   - data/base_cards.json       (베이스탐험카드)
#   - data/extra_cards.json      (1차_확정카드)
#   - data/combat_cards.json     (전투카드)
#   - data/monsters.json         (몬스터)
#   - data/prey.json             (사냥감)
#   - data/buildings.json        (빌딩, cost는 파서로 구조화)
#   - data/items.json            (Notion 아이템 DB 스냅샷)
#
# 사용법:
#   python3 scripts/fetch_data.py                      # API 우선, 실패 시 xlsx 폴백
#   python3 scripts/fetch_data.py --no-api             # xlsx export만 사용
#   python3 scripts/fetch_data.py --skip-sheet         # 캐시 xlsx로만 재생성
#
#   # 시트 쓰기 (API 전용):
#   python3 scripts/fetch_data.py --sheet-op=delete-row --tab=베이스탐험카드 --match-col=id --match-val=find_weapon
#   python3 scripts/fetch_data.py --sheet-op=append-row --tab=베이스탐험카드 --row-json='{"id":"rest",...}'
#
# Notion 아이템은 MCP/API 키 없이 로컬에서 동작해야 하므로, 이 스크립트는
# 이미 리포지토리에 커밋된 data/items.raw.json을 읽어 items.json을 생성한다.
# 최초 시드는 리사가 MCP로 수집해 items.raw.json에 저장해 두었고, 이후
# Notion에서 수정한 내용을 반영하고 싶을 때만 items.raw.json을 갱신하면 된다.
# (상세 절차는 README 하단 또는 '아이템 갱신 가이드' 문서 참고.)
#
# 디자인 메모 — 문자열 파서 방식 채택:
#   `cost`(빌딩), `attack_pattern`(몬스터) 등은 시트에서 자연어로 관리하도록 두고
#   이 파이프라인이 파싱해 구조화 필드를 추가한다. 실패해도 원문(_raw)을 병행
#   저장하므로 게임이 JSON을 읽을 때 폴백 가능.
# ────────────────────────────────────────────────────────────────────────

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.request
from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore
except ImportError:
    sys.stderr.write(
        "openpyxl이 필요합니다. 설치: pip3 install --user openpyxl\n"
    )
    sys.exit(1)

# gspread는 옵션. 없거나 native 의존성(cryptography/pyo3 등) 충돌로
# import 자체가 실패할 수 있다 → Exception 전부 잡아 xlsx 폴백으로 전환.
try:
    import gspread  # type: ignore
    from google.oauth2.service_account import Credentials  # type: ignore
    _HAS_GSPREAD = True
except BaseException as _gspread_err:
    # pyo3 PanicException·native 의존성 충돌은 Exception 하위가 아닐 수 있어
    # BaseException까지 확장. 단 사용자 인터럽트(Ctrl+C)·SystemExit은 그대로 전파.
    if isinstance(_gspread_err, (KeyboardInterrupt, SystemExit)):
        raise
    _HAS_GSPREAD = False


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
SHEET_ID = "1iS4Lmjx32w0Mu527foFtc8pQn3pw6APcQYnvcKdVM9o"
SHEET_XLSX_URL = (
    f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
)
SA_KEY_DEFAULT = ROOT / ".secrets" / "sheets-sa.json"
SA_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

# 시트 탭 이름 → 산출 JSON 경로 매핑
# '드롭테이블'은 별도 2-탭 구조(지역×카테고리% + 지역↔아이템 풀)이므로 여기 넣지 않고
# 아래 export_drop_table()에서 개별 처리.
SHEET_TABS = {
    "베이스탐험카드": "base_cards.json",
    "1차_확정카드": "extra_cards.json",
    "전투카드": "combat_cards.json",
    "몬스터": "monsters.json",
    "사냥감": "prey.json",
    "빌딩": "buildings.json",
    "특수카드": "special_cards.json",
}

# 드롭테이블 전용 탭 이름
DROP_TABLE_SHEET = "드롭테이블"
REGION_POOL_SHEET = "드롭풀"

# 아이템마스터 탭 (D-30): 아이템 SSOT 시트 이관
# Notion DB는 레거시(서사 참조용)로 남기고, 런타임은 여기서만 읽는다.
ITEM_MASTER_SHEET = "아이템마스터"
WEAPON_MASTER_SHEET = "무기"

# 조합레시피 탭 (D-30): 이원 재료 조합(같은 재료 아닌 것).
# 헤더: ingredient_a, ingredient_b, result, note
# 같은 재료 2개 머지는 아이템마스터의 merge_result 컬럼으로 표현.
COMBO_RECIPE_SHEET = "조합레시피"


# ─── 파서 ───────────────────────────────────────────────────────────────


COST_RE = re.compile(r"([가-힣A-Za-z_][가-힣A-Za-z0-9_ ]*?)\s*x\s*(\d+)")


def parse_cost(raw: str) -> list[dict[str, Any]]:
    """'목재 x10, 끈 x5' → [{material:'목재', count:10}, ...]"""
    if not raw:
        return []
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    out: list[dict[str, Any]] = []
    for part in parts:
        m = COST_RE.search(part)
        if not m:
            continue
        material = m.group(1).strip()
        count = int(m.group(2))
        out.append({"material": material, "count": count})
    return out


PATTERN_RE = re.compile(r"([가-힣A-Za-z_][가-힣A-Za-z0-9_ ]*?)\s*\((-?\d+)\)")


def parse_attack_pattern(raw: str) -> list[dict[str, Any]]:
    """'으르렁(0) → 물기(2) → 발톱(3)' → [{move:'으르렁', damage:0}, ...]"""
    if not raw:
        return []
    phases = re.split(r"→|->|→", str(raw))
    out: list[dict[str, Any]] = []
    for phase in phases:
        m = PATTERN_RE.search(phase)
        if not m:
            continue
        move = m.group(1).strip()
        dmg = int(m.group(2))
        out.append({"move": move, "damage": dmg})
    return out


# ─── 유틸 ───────────────────────────────────────────────────────────────


def normalize_cell(v: Any) -> Any:
    """xlsx 셀 값을 JSON 친화적으로 정규화.
    float인데 소수점이 0이면 int로, None은 None으로 유지."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def sheet_to_rows(ws) -> list[dict[str, Any]]:
    """첫 행을 키로 쓰는 dict 리스트로 변환."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [normalize_cell(h) for h in header]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue
        obj = {}
        for key, val in zip(header, row):
            if key is None:
                continue
            obj[str(key)] = normalize_cell(val)
        out.append(obj)
    return out


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")


# ─── Google Sheets API (gspread) ────────────────────────────────────────
#
# D-25 전환: 읽기 + 쓰기를 API로 일원화. xlsx export는 API 실패 시 폴백.
# 서비스 계정 키는 커밋하지 않으며 기본 경로는 .secrets/sheets-sa.json.
# 환경변수 GOOGLE_SHEETS_SA_KEY로 절대경로를 덮어쓸 수 있다.


def sa_key_path() -> Path:
    env = os.environ.get("GOOGLE_SHEETS_SA_KEY")
    if env:
        return Path(env).expanduser()
    return SA_KEY_DEFAULT


def open_spreadsheet() -> Any | None:
    """gspread Spreadsheet 핸들 반환. 의존성/키 없으면 None."""
    if not _HAS_GSPREAD:
        print("[api] gspread 미설치 — xlsx 폴백")
        return None
    key = sa_key_path()
    if not key.exists():
        print(f"[api] 서비스 계정 키 없음({key}) — xlsx 폴백")
        return None
    try:
        creds = Credentials.from_service_account_file(str(key), scopes=SA_SCOPES)
        gc = gspread.authorize(creds)
        return gc.open_by_key(SHEET_ID)
    except Exception as e:  # noqa: BLE001
        print(f"[api] 시트 열기 실패: {e} — xlsx 폴백")
        return None


def _ws_rows(ws) -> list[dict[str, Any]]:
    """gspread Worksheet → dict row 리스트. 첫 행을 헤더로 사용."""
    values = ws.get_all_values()
    if not values:
        return []
    header = [normalize_cell(h) for h in values[0]]
    out: list[dict[str, Any]] = []
    for row in values[1:]:
        if not any((c or "").strip() for c in row):
            continue
        obj: dict[str, Any] = {}
        for key, val in zip(header, row):
            if key is None or key == "":
                continue
            v = normalize_cell(val)
            # gspread는 모두 str로 주므로 숫자처럼 보이면 변환
            if isinstance(v, str):
                if re.fullmatch(r"-?\d+", v):
                    v = int(v)
                elif re.fullmatch(r"-?\d+\.\d+", v):
                    fv = float(v)
                    v = int(fv) if fv.is_integer() else fv
            obj[str(key)] = v
        out.append(obj)
    return out


def export_sheets_via_api(sh) -> bool:
    """API로 모든 탭을 읽어 JSON 덤프. 성공 True."""
    names = {w.title for w in sh.worksheets()}
    for tab_name, out_file in SHEET_TABS.items():
        if tab_name not in names:
            print(f"[warn] 탭 '{tab_name}' 이 시트에 없음 — 스킵")
            continue
        ws = sh.worksheet(tab_name)
        rows = _ws_rows(ws)

        if tab_name == "빌딩":
            for row in rows:
                raw_cost = row.get("cost", "")
                row["cost_raw"] = raw_cost
                row["cost"] = parse_cost(raw_cost or "")

        if tab_name == "몬스터":
            for row in rows:
                raw = row.get("attack_pattern", "")
                row["attack_pattern_raw"] = raw
                row["attack_pattern"] = parse_attack_pattern(raw or "")

        out_path = DATA_DIR / out_file
        write_json(out_path, rows)
        print(f"[api] {tab_name:12s} → {out_path.name}  ({len(rows)} rows)")
    return True


def _parse_weights_csv(raw: Any) -> list[int]:
    """가중치 CSV(\"80, 20\") → [80, 20]. 숫자 아닌 토큰은 0으로."""
    if raw is None:
        return []
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    out: list[int] = []
    for p in parts:
        try:
            out.append(int(float(p)))
        except ValueError:
            out.append(0)
    return out


def export_drop_table_via_api(sh) -> None:
    names = {w.title for w in sh.worksheets()}
    regions: dict[str, dict[str, int]] = {}
    pool: dict[str, dict[str, list[str]]] = {}
    weights: dict[str, dict[str, list[int]]] = {}

    if DROP_TABLE_SHEET in names:
        for row in _ws_rows(sh.worksheet(DROP_TABLE_SHEET)):
            region = row.get("region") or row.get("지역")
            if not region:
                continue
            regions[str(region)] = {
                "env": int(row.get("env") or 0),
                "food": int(row.get("food") or 0),
                "none": int(row.get("none") or 0),
            }
    else:
        print(f"[warn] 탭 '{DROP_TABLE_SHEET}' 이 시트에 없음 — 스킵")

    if REGION_POOL_SHEET in names:
        for row in _ws_rows(sh.worksheet(REGION_POOL_SHEET)):
            region = row.get("region") or row.get("지역")
            category = row.get("category") or row.get("카테고리")
            items_raw = row.get("items") or row.get("아이템") or ""
            if not region or not category:
                continue
            items = [s.strip() for s in str(items_raw).split(",") if s.strip()]
            pool.setdefault(str(region), {}).setdefault(str(category), []).extend(items)

            # D-33: weights 컬럼이 있으면 가중치 맵에 반영.
            # items와 길이가 같아야 의미 있음 — 다를 땐 런타임에서 균등 추첨으로 폴백.
            weights_raw = row.get("weights") or ""
            parsed_w = _parse_weights_csv(weights_raw)
            if parsed_w:
                weights.setdefault(str(region), {}).setdefault(str(category), []).extend(parsed_w)
    else:
        print(f"[warn] 탭 '{REGION_POOL_SHEET}' 이 시트에 없음 — 스킵")

    if not regions and not pool:
        print("[info] drop_table.json 생성 스킵 (시트 탭 없음)")
        return
    payload: dict[str, Any] = {"regions": regions, "pool": pool}
    if weights:
        payload["weights"] = weights
    out_path = DATA_DIR / "drop_table.json"
    write_json(out_path, payload)
    print(f"[api] drop_table  → {out_path.name}  ({len(regions)} regions)")


# ─── 시트 쓰기 헬퍼 (API 전용) ────────────────────────────────────────
#
# 탭 생성 / row 추가 / row 삭제 / 셀 편집. 모두 서비스 계정 권한 필요(편집자).


def sheet_ensure_tab(sh, tab: str, headers: list[str]) -> Any:
    """탭이 없으면 생성하고 헤더 세팅. Worksheet 반환."""
    try:
        ws = sh.worksheet(tab)
    except gspread.WorksheetNotFound:  # type: ignore[attr-defined]
        ws = sh.add_worksheet(title=tab, rows=100, cols=max(10, len(headers)))
        ws.update("A1", [headers])
        print(f"[write] 탭 생성 '{tab}' with {len(headers)} cols")
    return ws


def sheet_append_row(sh, tab: str, row_obj: dict[str, Any]) -> None:
    """탭의 첫 행을 헤더로 보고 row_obj를 매핑해 append."""
    ws = sh.worksheet(tab)
    header = ws.row_values(1)
    if not header:
        raise RuntimeError(f"탭 '{tab}' 헤더가 비어있음 — 먼저 헤더를 세팅하세요")
    values = [row_obj.get(h, "") for h in header]
    # None은 빈 문자열
    values = ["" if v is None else v for v in values]
    ws.append_row(values, value_input_option="USER_ENTERED")
    print(f"[write] append_row({tab}): {row_obj}")


def sheet_delete_row(sh, tab: str, match_col: str, match_val: str) -> int:
    """match_col == match_val인 첫 행을 삭제. 삭제한 행 번호 반환(없으면 -1)."""
    ws = sh.worksheet(tab)
    header = ws.row_values(1)
    if match_col not in header:
        raise RuntimeError(f"탭 '{tab}'에 컬럼 '{match_col}' 없음")
    col_idx = header.index(match_col) + 1  # gspread는 1-indexed
    col_values = ws.col_values(col_idx)
    # col_values[0]은 헤더, 1부터가 데이터
    for i, v in enumerate(col_values[1:], start=2):
        if str(v).strip() == str(match_val).strip():
            ws.delete_rows(i)
            print(f"[write] delete_row({tab}, {match_col}={match_val}) → row {i}")
            return i
    print(f"[write] delete_row: {match_col}={match_val} 없음 — no-op")
    return -1


def sheet_update_cell(sh, tab: str, match_col: str, match_val: str, set_col: str, new_val: Any) -> None:
    """match 조건 행의 set_col을 new_val로 변경."""
    ws = sh.worksheet(tab)
    header = ws.row_values(1)
    if match_col not in header or set_col not in header:
        raise RuntimeError(f"탭 '{tab}' 컬럼 누락: {match_col}/{set_col}")
    m_idx = header.index(match_col) + 1
    s_idx = header.index(set_col) + 1
    col_values = ws.col_values(m_idx)
    for i, v in enumerate(col_values[1:], start=2):
        if str(v).strip() == str(match_val).strip():
            ws.update_cell(i, s_idx, new_val)
            print(f"[write] update_cell({tab}, row{i}, {set_col}={new_val})")
            return
    raise RuntimeError(f"match 행 없음: {match_col}={match_val}")


# ─── 엔드포인트 ─────────────────────────────────────────────────────────


def fetch_sheet(xlsx_url: str, out_path: Path) -> None:
    print(f"[fetch] downloading xlsx from google sheets...")
    req = urllib.request.Request(
        xlsx_url,
        headers={"User-Agent": "ttd-fetch-data/1.0"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(data)
    print(f"[fetch] saved {len(data)} bytes → {out_path}")


def export_sheets(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    existing = set(wb.sheetnames)
    for tab_name, out_file in SHEET_TABS.items():
        if tab_name not in existing:
            print(f"[warn] 탭 '{tab_name}' 이 시트에 없음 — 스킵")
            continue
        rows = sheet_to_rows(wb[tab_name])

        # 빌딩: cost 파서 + _raw
        if tab_name == "빌딩":
            for row in rows:
                raw_cost = row.get("cost", "")
                row["cost_raw"] = raw_cost
                row["cost"] = parse_cost(raw_cost or "")

        # 몬스터: attack_pattern 파서 + _raw
        if tab_name == "몬스터":
            for row in rows:
                raw = row.get("attack_pattern", "")
                row["attack_pattern_raw"] = raw
                row["attack_pattern"] = parse_attack_pattern(raw or "")

        out_path = DATA_DIR / out_file
        write_json(out_path, rows)
        print(f"[export] {tab_name:12s} → {out_path.name}  ({len(rows)} rows)")


# ─── 드롭테이블: 시트 2탭 → drop_table.json ───────────────────────────
#
# 기존 dropTable.js의 DROP_TABLE / REGION_ITEM_POOL 하드코딩 값을
# 시트로 옮기기 위한 파서. 시트에 아래 두 탭이 있어야 한다.
#
#   탭 '드롭테이블' (지역별 카테고리 확률)
#     region | env | food | none
#     숲     | 50  | 25   | 25
#     ...
#
#   탭 '드롭풀' (지역×카테고리 → 아이템 타입 CSV)
#     region | category | items
#     숲     | env      | branch
#     숲     | food     | mushroom
#     덤불   | env      | stem
#     ...
#
# 산출물: data/drop_table.json
#   {
#     "regions": { "숲": { "env":50, "food":25, "none":25 }, ... },
#     "pool":    { "숲": { "env":["branch"], "food":["mushroom"] }, ... }
#   }


def export_drop_table(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    existing = set(wb.sheetnames)

    regions: dict[str, dict[str, int]] = {}
    pool: dict[str, dict[str, list[str]]] = {}
    weights: dict[str, dict[str, list[int]]] = {}

    # 1) 확률 탭
    if DROP_TABLE_SHEET in existing:
        for row in sheet_to_rows(wb[DROP_TABLE_SHEET]):
            region = row.get("region") or row.get("지역")
            if not region:
                continue
            regions[str(region)] = {
                "env": int(row.get("env") or 0),
                "food": int(row.get("food") or 0),
                "none": int(row.get("none") or 0),
            }
    else:
        print(f"[warn] 탭 '{DROP_TABLE_SHEET}' 이 시트에 없음 — 스킵")

    # 2) 풀 탭
    if REGION_POOL_SHEET in existing:
        for row in sheet_to_rows(wb[REGION_POOL_SHEET]):
            region = row.get("region") or row.get("지역")
            category = row.get("category") or row.get("카테고리")
            items_raw = row.get("items") or row.get("아이템") or ""
            if not region or not category:
                continue
            items = [s.strip() for s in str(items_raw).split(",") if s.strip()]
            pool.setdefault(str(region), {}).setdefault(str(category), []).extend(items)

            # D-33: weights 컬럼 — CSV 형식. 없으면 균등 추첨 폴백.
            weights_raw = row.get("weights") or ""
            parsed_w = _parse_weights_csv(weights_raw)
            if parsed_w:
                weights.setdefault(str(region), {}).setdefault(str(category), []).extend(parsed_w)
    else:
        print(f"[warn] 탭 '{REGION_POOL_SHEET}' 이 시트에 없음 — 스킵")

    if not regions and not pool:
        # 시트에 아직 탭이 없으면 drop_table.json을 만들지 않는다 (기존 dropTable.js 유지).
        print("[info] drop_table.json 생성 스킵 (시트 탭 없음)")
        return

    payload: dict[str, Any] = {"regions": regions, "pool": pool}
    if weights:
        payload["weights"] = weights
    out_path = DATA_DIR / "drop_table.json"
    write_json(out_path, payload)
    print(f"[export] drop_table  → {out_path.name}  ({len(regions)} regions, {sum(len(v) for v in pool.values())} pool entries)")


# ─── Notion 아이템: raw → items.json ───────────────────────────────────


def parse_effect_dsl(raw: str) -> dict[str, Any]:
    """'사용 효과' DSL을 게임 런타임이 먹기 좋은 구조로 파싱.

    DSL은 effectParser.js의 문법과 동일:
        'hunger+1'                  → 스탯 +1
        'spawn_card:throw'          → 전투 카드 추가
        'hunger+1;health+1'         → 체인
        ''/'미정'/'효과 없음'/'-'   → usable=False

    반환: { usable: bool, actions: [...], raw: str }
    런타임 JS 파서와 입력 규약이 일치하므로 둘 중 어느 쪽을 써도 된다.
    """
    s = str(raw or "").strip()
    if not s or s in {"미정", "효과 없음", "효과없음", "tbd", "TBD", "-"}:
        return {"usable": False, "actions": [], "raw": s}

    tokens = [t.strip() for t in s.split(";") if t.strip()]
    actions: list[dict[str, Any]] = []
    for tok in tokens:
        m_spawn = re.fullmatch(r"spawn_card\s*:\s*([A-Za-z0-9_\-]+)", tok)
        if m_spawn:
            actions.append({"type": "spawn_card", "cardId": m_spawn.group(1)})
            continue
        m_stat = re.fullmatch(r"([a-zA-Z_]+)\s*([+\-])\s*(\d+)", tok)
        if m_stat:
            delta = int(m_stat.group(3))
            if m_stat.group(2) == "-":
                delta = -delta
            actions.append({"type": "stat", "stat": m_stat.group(1).lower(), "delta": delta})
            continue
        # 인식 못한 토큰은 조용히 무시 (UI에서 raw 노출은 유지)

    return {"usable": len(actions) > 0, "actions": actions, "raw": s}


# ─── id 매핑 (Notion 한글 이름 ↔ 런타임 영문 key) ──────────────────────
#
# D-24: 머지·조합 통합 시스템에서 런타임은 영문 id(`stone`, `branch` …)로
# 아이템을 다루고, Notion DB는 한글 title(`돌맹이`, `나뭇가지` …)을 쓴다.
# 이 테이블이 두 축을 잇는 단일 진실의 원천. `inventory.js::ITEMS` 키와 정확히 1:1.
# 신규 아이템 추가 시 양쪽에 함께 등록.

ITEM_NAME_TO_ID: dict[str, str] = {
    # 1단계 (파밍)
    "돌맹이": "stone",
    "나뭇가지": "branch",
    "줄기": "stem",           # D-49 (2026-04-24): '질긴줄기' → '줄기' 개명
    "질긴줄기": "stem",       # 레거시 동의어 (구 조합법 raw 호환)
    "버섯": "mushroom",
    "산딸기": "berry",
    # 2단계 (조합)
    "목재": "wood",
    "끈": "plant_fiber",            # D-49 (2026-04-24): '식물 섬유(끈)' → '끈' 개명
    "식물 섬유(끈)": "plant_fiber", # 레거시 동의어
    "식물 섬유": "plant_fiber",     # 레거시 동의어
    "깨끗한 천": "clean_cloth",
    "붕대": "bandage",
}


def lookup_item_id(name_or_url: str, by_url: dict[str, str]) -> str | None:
    """이름 또는 Notion URL로 런타임 id를 찾는다.
    relation이 URL 배열로 오므로 by_url 도움이 필요하다."""
    if not name_or_url:
        return None
    s = str(name_or_url).strip()
    # URL 형태면 by_url에서 역추적
    if s.startswith("http"):
        return by_url.get(s)
    # 이름이면 직접
    return ITEM_NAME_TO_ID.get(s)


# ─── 조합법 DSL 파서 ───────────────────────────────────────────────────
#
# 예시 입력: "나뭇가지 + 나뭇가지 → 목재 / 목재 + 목재 → 튼튼한 목재(고급)"
# 각 slash-분절에서 `재료 + 재료 [+ 재료…] → 결과물` 패턴을 추출.
# "x3"·"x5" 수량 표기(`억센 풀 x3 → 끈 x1`)는 이번 버전에서 **무시**하고
# 기본 N=2 머지/2종 조합만 다룬다 — 3종 이상 및 수량 파싱은 후속 확장.
#
# 반환: (ingredients_names: list[str], result_name: str) 튜플 리스트
#
# 괄호 주석("(고급)", "(치료소 1단계 이상)")은 결과물에서 벗겨낸다.

COMBO_SEGMENT_RE = re.compile(
    r"([^/]+?)\s*(?:→|->)\s*([^/]+?)(?:\s*/|\s*$)"
)
PAREN_TRAIL_RE = re.compile(r"\s*\([^)]*\)\s*$")
QUANTITY_RE = re.compile(r"\s*x\s*\d+\s*$", re.IGNORECASE)


def _strip_item_name(s: str) -> str:
    """'튼튼한 목재(고급)' → '튼튼한 목재', '끈 x1' → '끈'"""
    if not s:
        return ""
    s = str(s).strip()
    s = PAREN_TRAIL_RE.sub("", s)
    s = QUANTITY_RE.sub("", s)
    return s.strip()


def parse_combo_recipe(raw: str) -> list[tuple[list[str], str]]:
    """Notion `조합법` 텍스트를 (ingredients, result) 튜플 리스트로 파싱."""
    if not raw:
        return []
    out: list[tuple[list[str], str]] = []
    # slash 단위로 자른 뒤 각 분절에서 → 좌우를 추출
    for segment in str(raw).split("/"):
        seg = segment.strip()
        if "→" not in seg and "->" not in seg:
            continue
        # 정규식보단 단순 split이 안전
        left, _, right = seg.partition("→")
        if not right:
            left, _, right = seg.partition("->")
        if not left or not right:
            continue
        ingredients = [_strip_item_name(x) for x in left.split("+")]
        ingredients = [x for x in ingredients if x]
        result = _strip_item_name(right)
        if not ingredients or not result:
            continue
        out.append((ingredients, result))
    return out


def transform_notion_items(raw: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """MCP notion-fetch 응답의 properties dict 배열을 게임이 쓰기 좋은 형태로 가공.

    Notion 체크박스는 '__YES__'/'__NO__' 문자열로 오고, 멀티셀렉트는 list[str].
    여기서 Boolean/lowercase 타입으로 정리하되, 원본 구조는 크게 깨지 않는다.
    '사용 효과' rich_text는 파싱해 effect 구조를 덧붙인다 (원본도 유지).
    D-24: 머지·조합 통합 시스템용 필드 추가.
    """
    # URL → id 역인덱스 (relation 해석용)
    by_url: dict[str, str] = {}
    for props in raw:
        url = props.get("url")
        name = props.get("이름") or props.get("name")
        if url and name:
            rid = ITEM_NAME_TO_ID.get(name)
            if rid:
                by_url[url] = rid

    out = []
    for props in raw:
        item = dict(props)

        # boolean 정규화
        for bkey in ("머지 가능", "일회용"):
            v = item.get(bkey)
            if v == "__YES__":
                item[bkey] = True
            elif v == "__NO__":
                item[bkey] = False
            elif v is None:
                item[bkey] = False

        # 이름을 id-ish 키 'name'으로도 노출 (영문 접근용)
        if "이름" in item and "name" not in item:
            item["name"] = item["이름"]

        # D-24: 런타임 id (영문 key) 노출
        name = item.get("이름") or item.get("name")
        rid = ITEM_NAME_TO_ID.get(name) if name else None
        if rid:
            item["id"] = rid

        # D-24: 머지 결과물 id (자기참조 relation[0] → 이름 → id)
        # Notion의 '결과물' relation은 자기참조 페이지 URL 배열. 첫 원소가 2단계 결과물.
        result_urls = item.get("결과물") or []
        merge_result_id: str | None = None
        if isinstance(result_urls, list) and result_urls:
            first = result_urls[0]
            merge_result_id = by_url.get(first)
        item["merge_result"] = merge_result_id
        # 편의: merge_enabled는 '머지 가능' + merge_result 유효 둘 다 요구
        item["merge_enabled"] = bool(item.get("머지 가능")) and merge_result_id is not None

        # 사용 효과 파싱 (Notion rich_text 원문 → 구조화)
        effect_raw = item.get("사용 효과", "")
        item["effect"] = parse_effect_dsl(effect_raw)

        out.append(item)

    # 아이템 ID(ITM-1 …) 순 정렬 시도
    def _key(it):
        raw = it.get("아이템 ID", "")
        m = re.search(r"(\d+)$", str(raw))
        return int(m.group(1)) if m else 9999

    out.sort(key=_key)
    return out


def build_combos_from_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """아이템들의 `조합법` 텍스트를 모두 파싱해 전역 combos 리스트 생성.

    D-24 정책:
    - 2종 이상 재료만 허용 (머지는 같은 재료 2개 = combos에도 등장 가능; 둘 다 지원).
    - 괄호 주석 · 수량 표기는 무시.
    - 재료/결과 이름이 ITEM_NAME_TO_ID에 없으면 스킵(경고).
    - 중복 레시피는 제거(ingredients sort + result 키).
    """
    seen: set[tuple[tuple[str, ...], str]] = set()
    combos: list[dict[str, Any]] = []
    for it in items:
        recipe_raw = it.get("조합법", "") or ""
        for ingredients, result in parse_combo_recipe(recipe_raw):
            ing_ids: list[str] = []
            bad = False
            for name in ingredients:
                rid = ITEM_NAME_TO_ID.get(name)
                if not rid:
                    print(f"[warn] 조합 재료 미매핑: '{name}' (from {it.get('이름')})")
                    bad = True
                    break
                ing_ids.append(rid)
            if bad:
                continue
            result_id = ITEM_NAME_TO_ID.get(result)
            if not result_id:
                print(f"[warn] 조합 결과 미매핑: '{result}' (from {it.get('이름')})")
                continue
            key = (tuple(sorted(ing_ids)), result_id)
            if key in seen:
                continue
            seen.add(key)
            combos.append({"ingredients": ing_ids, "result": result_id})
    return combos


# ─── 시트 '아이템마스터' → items.json / combos.json (D-30 SSOT 이관) ─────
#
# 스키마(헤더): id, name, category, material_type, grade, size, weight,
#              regions, mergeable, merge_result, disposable, effect,
#              summary, description, durability
#
# - mergeable/disposable: 'Y'/'N'
# - regions: 콤마 구분 CSV ('숲,동굴')
# - effect: DSL ('hunger+1', 'spawn_card:throw', 공란 = 사용불가)
# - merge_result: 같은 재료 2개 합쳤을 때 결과 id (Y이면 필수)
#
# 런타임 호환 필드(기존 inventory.js / effectParser.js가 읽던 것)를 함께 방출:
#   이름, 카테고리, 재료 타입, 사용 효과, 머지 가능, 일회용, 나오는 지역, 설명 텍스트, 효과 요약
#
# combos.json: 자기 자신 머지(branch+branch→wood)도 포함시킨다.
#   inventory.js::canMerge는 static ITEMS + TTD_DATA.ITEMS의 merge_result를 보지만,
#   lookupCombo도 같은 결과를 내주면 회귀 없이 "머지이자 조합"으로 양쪽 호환.


def _yn(v: Any) -> bool:
    s = str(v or "").strip().upper()
    return s in {"Y", "YES", "TRUE", "1"}


def _csv_list(v: Any) -> list[str]:
    if not v:
        return []
    return [s.strip() for s in str(v).split(",") if s.strip()]


def rows_to_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in rows:
        item_id = str(r.get("id") or "").strip()
        name = str(r.get("name") or "").strip()
        if not item_id or not name:
            continue
        mergeable = _yn(r.get("mergeable"))
        disposable = _yn(r.get("disposable"))
        merge_result = str(r.get("merge_result") or "").strip() or None
        effect_raw = str(r.get("effect") or "").strip()
        regions = _csv_list(r.get("regions"))

        item: dict[str, Any] = {
            # 런타임 영문 키 (inventory.js, dropTable.js가 읽는 것)
            "id": item_id,
            "name": name,
            # 런타임 호환: 한글 키 (effectParser.js matchesConsumeFilter, UI)
            "이름": name,
            "카테고리": str(r.get("category") or "").strip() or None,
            "재료 타입": str(r.get("material_type") or "").strip() or None,
            "아이템 등급": str(r.get("grade") or "").strip() or None,
            "가방칸수": str(r.get("size") or "").strip() or "1x1",
            "무게": r.get("weight"),
            "나오는 지역": regions,
            "머지 가능": mergeable,
            "일회용": disposable,
            "설명 텍스트": str(r.get("description") or "").strip(),
            "효과 요약": str(r.get("summary") or "").strip(),
            "사용 효과": effect_raw,
            # 머지 통합 시스템 (D-24 필드)
            "merge_result": merge_result,
            "merge_enabled": mergeable and bool(merge_result),
            # 내구도 (있을 때만)
        }
        durability = r.get("durability")
        if durability not in (None, "", 0):
            item["내구도"] = durability

        # 사용 효과 파싱 → { usable, actions, raw }
        item["effect"] = parse_effect_dsl(effect_raw)
        out.append(item)

    # grade → id 순 정렬 (2단계가 밑)
    def _key(it):
        g = it.get("아이템 등급") or ""
        g_num = 9
        m = re.search(r"(\d+)", g)
        if m:
            g_num = int(m.group(1))
        return (g_num, it.get("id", ""))
    out.sort(key=_key)
    return out


def build_combos_from_sheet(
    items: list[dict[str, Any]],
    extra_rows: list[dict[str, Any]] | None = None,
    extra_valid_ids: set[str] | None = None,
) -> list[dict[str, Any]]:
    """시트 `머지가능` + `merge_result`(같은 재료 머지) + `조합레시피` 탭(이원/삼원/사원 조합)을 합쳐 combos.json 생성.

    - 머지(같은 재료 2개): 아이템마스터의 merge_result 컬럼 — ingredients=[id, id]
    - 이원 조합: 조합레시피 탭 a/b → ingredients=[a, b]
    - 삼원/사원 조합(D-47): c/d 옵션 채움 → ingredients=[a, b, c(, d)]
    - 중복 제거(정렬한 ingredients + result 키).
    - extra_valid_ids: 아이템 외 유효 id(예: 무기). result 검증용 — wood+stone+plant_fiber=weapon_basic 같은 레시피 지원.
    """
    valid_ids = {it.get("id") for it in items if it.get("id")}
    if extra_valid_ids:
        valid_ids.update(extra_valid_ids)
    seen: set[tuple[tuple[str, ...], str]] = set()
    combos: list[dict[str, Any]] = []

    # 1) 같은 재료 머지
    for it in items:
        if not it.get("merge_enabled"):
            continue
        src = it.get("id")
        dst = it.get("merge_result")
        if not src or not dst:
            continue
        if dst not in valid_ids:
            print(f"[warn] merge_result '{dst}' 가 items에 없음 (from {src}) — 스킵")
            continue
        key = (tuple(sorted([src, src])), dst)
        if key in seen:
            continue
        seen.add(key)
        combos.append({"ingredients": [src, src], "result": dst})

    # 2) 이원·삼원·사원 조합 — a/b 필수, c/d 옵션.
    #    D-47: ingredient_d 추가 — 4재료 레시피(예: stone+stone+wood+plant_fiber=weapon_basic).
    for row in extra_rows or []:
        a = str(row.get("ingredient_a") or "").strip()
        b = str(row.get("ingredient_b") or "").strip()
        c = str(row.get("ingredient_c") or "").strip()
        d = str(row.get("ingredient_d") or "").strip()
        r = str(row.get("result") or "").strip()
        if not (a and b and r):
            continue
        ingredients = [a, b]
        if c:
            ingredients.append(c)
        if d:
            ingredients.append(d)
        missing = [x for x in ingredients + [r] if x not in valid_ids]
        if missing:
            print(f"[warn] 조합레시피 미매핑: {'+'.join(ingredients)}→{r} (미정의 id: {missing}) — 스킵")
            continue
        key = (tuple(sorted(ingredients)), r)
        if key in seen:
            continue
        seen.add(key)
        combos.append({"ingredients": ingredients, "result": r})

    return combos


def rows_to_weapons(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """D-47 무기 탭 파서. 스키마(헤더):
    id, name, size, weight, grade, durability, attack, description, summary
    출력은 items.json 스타일과 키 네이밍을 맞춤(한글 키 병행) — inventory.js
    resolveDef가 동일 경로에서 덮어쓰기로 사용할 수 있도록.
    """
    out: list[dict[str, Any]] = []
    for r in rows:
        wid = str(r.get("id") or "").strip()
        name = str(r.get("name") or "").strip()
        if not wid or not name:
            continue
        weapon: dict[str, Any] = {
            "id": wid,
            "name": name,
            "이름": name,
            "카테고리": "무기",
            "가방칸수": str(r.get("size") or "").strip() or "1x1",
            "무게": r.get("weight"),
            "아이템 등급": str(r.get("grade") or "").strip() or None,
            "내구도": r.get("durability"),
            "공격력": r.get("attack"),
            "설명 텍스트": str(r.get("description") or "").strip(),
            "효과 요약": str(r.get("summary") or "").strip(),
        }
        out.append(weapon)
    return out


def export_items_from_sheet(sh) -> bool:
    """API로 아이템마스터 + 무기 + 조합레시피 탭을 읽어 items/weapons/combos 생성."""
    names = {w.title for w in sh.worksheets()}
    if ITEM_MASTER_SHEET not in names:
        print(f"[warn] 탭 '{ITEM_MASTER_SHEET}' 없음 — items 생성 스킵")
        return False
    rows = _ws_rows(sh.worksheet(ITEM_MASTER_SHEET))
    items = rows_to_items(rows)
    out_path = DATA_DIR / "items.json"
    write_json(out_path, items)
    print(f"[api] items        → {out_path.name}  ({len(items)} items)")

    # D-47 무기 탭
    weapons: list[dict[str, Any]] = []
    if WEAPON_MASTER_SHEET in names:
        w_rows = _ws_rows(sh.worksheet(WEAPON_MASTER_SHEET))
        weapons = rows_to_weapons(w_rows)
    weapons_path = DATA_DIR / "weapons.json"
    write_json(weapons_path, weapons)
    print(f"[api] weapons      → {weapons_path.name}  ({len(weapons)} weapons)")

    combo_rows: list[dict[str, Any]] = []
    if COMBO_RECIPE_SHEET in names:
        combo_rows = _ws_rows(sh.worksheet(COMBO_RECIPE_SHEET))
    extra_ids = {w["id"] for w in weapons if w.get("id")}
    combos = build_combos_from_sheet(items, combo_rows, extra_valid_ids=extra_ids)
    combos_path = DATA_DIR / "combos.json"
    write_json(combos_path, combos)
    print(f"[api] combos       → {combos_path.name}  ({len(combos)} recipes)")
    return True


def export_items_from_xlsx(xlsx_path: Path) -> bool:
    """xlsx 폴백 — 아이템마스터 + 무기 + 조합레시피 탭에서 읽기."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if ITEM_MASTER_SHEET not in wb.sheetnames:
        print(f"[warn] xlsx에 '{ITEM_MASTER_SHEET}' 탭 없음 — items 생성 스킵")
        return False
    rows = sheet_to_rows(wb[ITEM_MASTER_SHEET])
    items = rows_to_items(rows)
    out_path = DATA_DIR / "items.json"
    write_json(out_path, items)
    print(f"[export] items       → {out_path.name}  ({len(items)} items)")

    weapons: list[dict[str, Any]] = []
    if WEAPON_MASTER_SHEET in wb.sheetnames:
        w_rows = sheet_to_rows(wb[WEAPON_MASTER_SHEET])
        weapons = rows_to_weapons(w_rows)
    weapons_path = DATA_DIR / "weapons.json"
    write_json(weapons_path, weapons)
    print(f"[export] weapons     → {weapons_path.name}  ({len(weapons)} weapons)")

    combo_rows: list[dict[str, Any]] = []
    if COMBO_RECIPE_SHEET in wb.sheetnames:
        combo_rows = sheet_to_rows(wb[COMBO_RECIPE_SHEET])
    extra_ids = {w["id"] for w in weapons if w.get("id")}
    combos = build_combos_from_sheet(items, combo_rows, extra_valid_ids=extra_ids)
    combos_path = DATA_DIR / "combos.json"
    write_json(combos_path, combos)
    print(f"[export] combos      → {combos_path.name}  ({len(combos)} recipes)")
    return True


def export_items() -> None:
    """레거시 경로 — 아이템 SSOT가 시트로 이관되기 전 Notion raw를 읽었던 함수.
    D-30부터는 쓰이지 않는다. items.raw.json은 레거시 보관만 하고 읽지 않는다.
    """
    print("[info] export_items(legacy) — Notion raw 경로는 D-30부터 비활성화됨")


# ─── main ──────────────────────────────────────────────────────────────


def run_sheet_op(args) -> int:
    """--sheet-op 처리. API 필수."""
    sh = open_spreadsheet()
    if sh is None:
        print("[error] sheet-op은 API 필요 — 서비스 계정 키 세팅 후 재시도")
        return 2
    op = args.sheet_op
    tab = args.tab
    if not tab and op != "ensure-tab":
        print("[error] --tab 필요")
        return 2

    if op == "delete-row":
        if not (args.match_col and args.match_val):
            print("[error] delete-row은 --match-col / --match-val 필요")
            return 2
        sheet_delete_row(sh, tab, args.match_col, args.match_val)
    elif op == "append-row":
        if not args.row_json:
            print("[error] append-row은 --row-json 필요")
            return 2
        row = json.loads(args.row_json)
        sheet_append_row(sh, tab, row)
    elif op == "update-cell":
        if not (args.match_col and args.match_val and args.set_col):
            print("[error] update-cell은 --match-col/--match-val/--set-col 필요")
            return 2
        sheet_update_cell(sh, tab, args.match_col, args.match_val, args.set_col, args.set_val)
    elif op == "ensure-tab":
        if not (tab and args.headers):
            print("[error] ensure-tab은 --tab / --headers 필요")
            return 2
        headers = [h.strip() for h in args.headers.split(",") if h.strip()]
        sheet_ensure_tab(sh, tab, headers)
    else:
        print(f"[error] 알 수 없는 sheet-op: {op}")
        return 2
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--skip-sheet",
        action="store_true",
        help="시트 다운로드 건너뛰기 (오프라인 — 마지막 xlsx 캐시 사용)",
    )
    ap.add_argument(
        "--no-items",
        action="store_true",
        help="items.json 생성 건너뛰기",
    )
    ap.add_argument(
        "--no-api",
        action="store_true",
        help="Sheets API 사용 안 함. xlsx export만 사용",
    )
    # 시트 쓰기 명령 (있으면 읽기 파이프라인 대신 이것만 실행)
    ap.add_argument("--sheet-op", choices=["delete-row", "append-row", "update-cell", "ensure-tab"])
    ap.add_argument("--tab")
    ap.add_argument("--match-col")
    ap.add_argument("--match-val")
    ap.add_argument("--set-col")
    ap.add_argument("--set-val")
    ap.add_argument("--row-json", help="JSON 문자열 (append-row용)")
    ap.add_argument("--headers", help="콤마 구분 (ensure-tab용)")
    args = ap.parse_args()

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # 쓰기 명령은 읽기 파이프라인과 분리
    if args.sheet_op:
        return run_sheet_op(args)

    # 읽기 파이프라인: API 우선, 실패 시 xlsx 폴백
    used_api = False
    if not args.no_api:
        sh = open_spreadsheet()
        if sh is not None:
            try:
                export_sheets_via_api(sh)
                export_drop_table_via_api(sh)
                if not args.no_items:
                    export_items_from_sheet(sh)
                used_api = True
                print("[api] 시트 읽기 완료 (Sheets API)")
            except Exception as e:  # noqa: BLE001
                print(f"[api] 읽기 중 예외: {e} — xlsx 폴백 시도")

    if not used_api:
        xlsx_cache = DATA_DIR / ".sheet_cache.xlsx"
        if not args.skip_sheet:
            try:
                fetch_sheet(SHEET_XLSX_URL, xlsx_cache)
            except Exception as e:
                print(f"[error] 시트 다운로드 실패: {e}")
                if not xlsx_cache.exists():
                    return 1
                print("[warn] 기존 캐시로 진행")
        else:
            if not xlsx_cache.exists():
                print(f"[error] --skip-sheet 지정됐지만 캐시({xlsx_cache})가 없음")
                return 1
        export_sheets(xlsx_cache)
        export_drop_table(xlsx_cache)
        if not args.no_items:
            export_items_from_xlsx(xlsx_cache)

    # 브라우저 로더용 data.js 생성 (file:// 환경에서 fetch 없이 사용)
    export_data_js()

    print("[ok] 완료")
    return 0


# ─── 브라우저 로더(data.js) ─────────────────────────────────────────────


BROWSER_BUNDLE_KEYS = {
    "base_cards.json": "BASE_CARDS",
    "extra_cards.json": "EXTRA_CARDS",
    "combat_cards.json": "COMBAT_CARDS",
    "monsters.json": "MONSTERS",
    "prey.json": "PREY",
    "special_cards.json": "SPECIAL_CARDS",
    "buildings.json": "BUILDINGS",
    "items.json": "ITEMS",
    "weapons.json": "WEAPONS",  # D-47: 무기 테이블 (별도 탭 '무기')
    "drop_table.json": "DROP_TABLE",
    "combos.json": "COMBOS",  # D-24: 머지·조합 통합 시스템
}


def export_data_js() -> None:
    """data.js — 전역 객체 `TTD_DATA`를 심어 <script>로 로드 가능하게."""
    bundle: dict[str, Any] = {}
    for filename, key in BROWSER_BUNDLE_KEYS.items():
        path = DATA_DIR / filename
        if not path.exists():
            bundle[key] = []
            continue
        bundle[key] = json.loads(path.read_text(encoding="utf-8"))

    js_lines = [
        "// data/data.js — AUTO-GENERATED by scripts/fetch_data.py",
        "// 원본 편집 금지. 시트/Notion을 수정한 뒤 `make data`로 재생성할 것.",
        "window.TTD_DATA = " + json.dumps(bundle, ensure_ascii=False, indent=2) + ";",
        "",
    ]
    out = DATA_DIR / "data.js"
    out.write_text("\n".join(js_lines), encoding="utf-8")
    print(f"[export] browser bundle → {out.name}")


if __name__ == "__main__":
    raise SystemExit(main())
